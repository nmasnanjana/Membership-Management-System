"""
Enhanced export utilities for Excel and PDF exports
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from django.http import HttpResponse
from datetime import datetime
from .models import Member, MeetingInfo, MemberAttendance


def style_excel_worksheet(ws, title):
    """Apply consistent styling to Excel worksheet"""
    # Title row
    title_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Style header row
    for cell in ws[1]:
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def export_members_excel(members, filename=None, selected_columns=None):
    """Export members to Excel with enhanced formatting and column selection"""
    if filename is None:
        filename = f"Members_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    # Column mapping
    column_map = {
        'member_id': ('Member ID', lambda m: m.member_id),
        'initials': ('Initials', lambda m: m.member_initials),
        'first_name': ('First Name', lambda m: m.member_first_name),
        'last_name': ('Last Name', lambda m: m.member_last_name),
        'address': ('Address', lambda m: m.member_address),
        'dob': ('Date of Birth', lambda m: m.member_dob.strftime("%d/%m/%Y") if m.member_dob else ""),
        'telephone': ('Telephone', lambda m: m.member_tp_number),
        'account': ('Account Number', lambda m: m.member_acc_number or ""),
        'guardian': ('Guardian Name', lambda m: m.member_guardian_name),
        'role': ('Role', lambda m: m.get_member_role_display() if m.member_role else "No Role"),
        'status': ('Status', lambda m: "Active" if m.member_is_active else "Inactive"),
        'join_date': ('Join Date', lambda m: m.member_join_at.strftime("%d/%m/%Y") if m.member_join_at else "")
    }
    
    # Default columns if none selected
    if selected_columns is None:
        selected_columns = list(column_map.keys())
    
    # Filter to only include valid columns
    selected_columns = [col for col in selected_columns if col in column_map]
    if not selected_columns:
        selected_columns = list(column_map.keys())
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    
    # Build headers and data based on selected columns
    headers = [column_map[col][0] for col in selected_columns]
    ws.append(headers)
    
    for member in members:
        row = [column_map[col][1](member) for col in selected_columns]
        ws.append(row)
    
    style_excel_worksheet(ws, "Members")
    wb.save(response)
    return response


def export_members_pdf(members, filename=None, selected_columns=None):
    """Export members to PDF with column selection"""
    if filename is None:
        filename = f"Members_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    # Column mapping
    column_map = {
        'member_id': ('Member ID', lambda m: m.member_id),
        'initials': ('Initials', lambda m: m.member_initials),
        'first_name': ('First Name', lambda m: m.member_first_name),
        'last_name': ('Last Name', lambda m: m.member_last_name),
        'address': ('Address', lambda m: m.member_address),
        'dob': ('Date of Birth', lambda m: m.member_dob.strftime("%d/%m/%Y") if m.member_dob else ""),
        'telephone': ('Telephone', lambda m: m.member_tp_number),
        'account': ('Account Number', lambda m: m.member_acc_number or ""),
        'guardian': ('Guardian Name', lambda m: m.member_guardian_name),
        'role': ('Role', lambda m: m.get_member_role_display() if m.member_role else "No Role"),
        'status': ('Status', lambda m: "Active" if m.member_is_active else "Inactive"),
        'join_date': ('Join Date', lambda m: m.member_join_at.strftime("%d/%m/%Y") if m.member_join_at else "")
    }
    
    # Default columns if none selected (limit to 6 for PDF readability)
    if selected_columns is None:
        selected_columns = ['member_id', 'first_name', 'last_name', 'telephone', 'role', 'status']
    
    # Filter to only include valid columns
    selected_columns = [col for col in selected_columns if col in column_map]
    if not selected_columns:
        selected_columns = ['member_id', 'first_name', 'last_name', 'telephone', 'role', 'status']
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f2937'),
        spaceAfter=30,
        alignment=1  # Center
    )
    title = Paragraph("Members Export Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Build table data based on selected columns
    headers = [column_map[col][0] for col in selected_columns]
    data = [headers]
    
    for member in members:
        row = [column_map[col][1](member) for col in selected_columns]
        data.append(row)
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#374151')),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.whitesmoke),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#4b5563')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


def export_meetings_excel(meetings, filename=None):
    """Export meetings to Excel"""
    if filename is None:
        filename = f"Meetings_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Meetings"
    
    headers = ["Meeting ID", "Date", "Fee", "Created At"]
    ws.append(headers)
    
    for meeting in meetings:
        ws.append([
            meeting.meeting_id,
            meeting.meeting_date.strftime("%d/%m/%Y") if meeting.meeting_date else "",
            meeting.meeting_fee,
            meeting.meeting_created_at.strftime("%d/%m/%Y %H:%M") if meeting.meeting_created_at else ""
        ])
    
    style_excel_worksheet(ws, "Meetings")
    wb.save(response)
    return response


def export_attendance_excel(attendances, filename=None):
    """Export attendance records to Excel"""
    if filename is None:
        filename = f"Attendance_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    headers = ["Member ID", "Member Name", "Meeting Date", "Attendance", "Fee Paid"]
    ws.append(headers)
    
    for attendance in attendances.select_related('member_id', 'meeting_date'):
        member = attendance.member_id
        full_name = f"{member.member_initials} {member.member_first_name} {member.member_last_name}"
        meeting_date = attendance.meeting_date.meeting_date.strftime("%d/%m/%Y") if attendance.meeting_date.meeting_date else ""
        attendance_status = "Present" if attendance.attendance_status else "Absent"
        fee_status = "Paid" if attendance.attendance_fee_status else "Not Paid"
        
        ws.append([
            member.member_id,
            full_name,
            meeting_date,
            attendance_status,
            fee_status
        ])
    
    style_excel_worksheet(ws, "Attendance")
    wb.save(response)
    return response

