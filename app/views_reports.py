"""
Custom Reports Builder
"""
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Count, Sum, Q, Avg
from .models import Member, MeetingInfo, MemberAttendance, Payment
from .views import context_data
from .export_utils import export_members_excel, export_members_pdf, export_meetings_excel, export_attendance_excel
from datetime import datetime, date, timedelta


@login_required
def reports_builder(request):
    """Custom reports builder interface"""
    context = context_data(request)
    context['page_name'] = 'Custom Reports Builder'
    
    if request.method == 'POST':
        report_type = request.POST.get('report_type')
        date_from = request.POST.get('date_from', '')
        date_to = request.POST.get('date_to', '')
        filters = {}
        
        # Build filters based on report type
        if report_type == 'members':
            filters['is_active'] = request.POST.get('member_status', '')
            filters['role'] = request.POST.get('member_role', '')
            filters['join_date_from'] = request.POST.get('join_date_from', '')
            filters['join_date_to'] = request.POST.get('join_date_to', '')
        elif report_type == 'attendance':
            filters['member_id'] = request.POST.get('member_id', '')
            filters['meeting_id'] = request.POST.get('meeting_id', '')
            filters['attendance_status'] = request.POST.get('attendance_status', '')
            filters['fee_status'] = request.POST.get('fee_status', '')
            if date_from:
                filters['date_from'] = date_from
            if date_to:
                filters['date_to'] = date_to
        elif report_type == 'meetings':
            if date_from:
                filters['date_from'] = date_from
            if date_to:
                filters['date_to'] = date_to
            filters['fee_min'] = request.POST.get('fee_min', '')
            filters['fee_max'] = request.POST.get('fee_max', '')
        elif report_type == 'payments':
            filters['member_id'] = request.POST.get('member_id', '')
            filters['payment_method'] = request.POST.get('payment_method', '')
            if date_from:
                filters['date_from'] = date_from
            if date_to:
                filters['date_to'] = date_to
        
        export_format = request.POST.get('export_format', 'excel')
        
        # Generate report
        try:
            if report_type == 'members':
                from .search_utils import search_members
                members = search_members('', filters)
                if export_format == 'pdf':
                    return export_members_pdf(members)
                else:
                    return export_members_excel(members)
            
            elif report_type == 'attendance':
                from .search_utils import search_attendance
                attendances = search_attendance('', filters)
                return export_attendance_excel(attendances)
            
            elif report_type == 'meetings':
                from .search_utils import search_meetings
                meetings = search_meetings('', filters)
                return export_meetings_excel(meetings)
            
            elif report_type == 'payments':
                payments = Payment.objects.select_related('member', 'meeting').all()
                if filters.get('member_id'):
                    payments = payments.filter(member__member_id__icontains=filters['member_id'])
                if filters.get('payment_method'):
                    payments = payments.filter(payment_method=filters['payment_method'])
                if filters.get('date_from'):
                    payments = payments.filter(payment_date__date__gte=filters['date_from'])
                if filters.get('date_to'):
                    payments = payments.filter(payment_date__date__lte=filters['date_to'])
                
                # Export payments to Excel
                from openpyxl import Workbook
                from openpyxl.styles import Font, PatternFill, Alignment
                response = HttpResponse(content_type='application/ms-excel')
                response['Content-Disposition'] = f'attachment; filename="Payments_Report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                
                wb = Workbook()
                ws = wb.active
                ws.title = "Payments"
                
                headers = ["Payment ID", "Member ID", "Member Name", "Amount", "Method", "Meeting Date", "Receipt #", "Date"]
                ws.append(headers)
                
                for payment in payments:
                    ws.append([
                        payment.payment_id,
                        payment.member.member_id,
                        f"{payment.member.member_first_name} {payment.member.member_last_name}",
                        float(payment.amount),
                        payment.get_payment_method_display(),
                        payment.meeting.meeting_date.strftime("%d/%m/%Y") if payment.meeting else "",
                        payment.receipt_number or "",
                        payment.payment_date.strftime("%d/%m/%Y %H:%M")
                    ])
                
                # Style header
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.alignment = Alignment(horizontal='center')
                
                wb.save(response)
                return response
            
        except Exception as e:
            messages.error(request, f'Error generating report: {str(e)}')
            return redirect('reports_builder')
    
    # Get data for filters
    members = Member.objects.filter(member_is_active=True).order_by('member_first_name')[:100]
    meetings = MeetingInfo.objects.order_by('-meeting_date')[:50]
    
    context.update({
        'members': members,
        'meetings': meetings,
        'breadcrumb_items': [
            {'name': 'Dashboard', 'url': '/', 'icon': 'home'},
            {'name': 'Reports Builder', 'icon': 'file-alt'},
        ]
    })
    
    return render(request, 'reports/builder.html', context)


@login_required
def reports_quick_stats(request):
    """Quick statistics report"""
    context = context_data(request)
    context['page_name'] = 'Quick Statistics'
    
    # Get date range
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    if not date_from:
        date_from = (date.today() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not date_to:
        date_to = date.today().strftime('%Y-%m-%d')
    
    # Build queries
    members_query = Member.objects.all()
    meetings_query = MeetingInfo.objects.all()
    attendance_query = MemberAttendance.objects.all()
    payments_query = Payment.objects.all()
    
    if date_from:
        meetings_query = meetings_query.filter(meeting_date__gte=date_from)
        attendance_query = attendance_query.filter(meeting_date__meeting_date__gte=date_from)
        payments_query = payments_query.filter(payment_date__date__gte=date_from)
    
    if date_to:
        meetings_query = meetings_query.filter(meeting_date__lte=date_to)
        attendance_query = attendance_query.filter(meeting_date__meeting_date__lte=date_to)
        payments_query = payments_query.filter(payment_date__date__lte=date_to)
    
    # Calculate statistics
    stats = {
        'total_members': members_query.count(),
        'active_members': members_query.filter(member_is_active=True).count(),
        'total_meetings': meetings_query.count(),
        'total_attendance': attendance_query.count(),
        'present_count': attendance_query.filter(attendance_status=True).count(),
        'paid_count': attendance_query.filter(attendance_fee_status=True).count(),
        'total_payments': payments_query.aggregate(total=Sum('amount'))['total'] or 0,
        'payment_count': payments_query.count(),
    }
    
    if stats['total_attendance'] > 0:
        stats['attendance_rate'] = (stats['present_count'] / stats['total_attendance']) * 100
    else:
        stats['attendance_rate'] = 0
    
    context.update({
        'stats': stats,
        'date_from': date_from,
        'date_to': date_to,
        'breadcrumb_items': [
            {'name': 'Dashboard', 'url': '/', 'icon': 'home'},
            {'name': 'Reports', 'url': '/reports/builder/', 'icon': 'file-alt'},
            {'name': 'Quick Stats', 'icon': 'chart-bar'},
        ]
    })
    
    return render(request, 'reports/quick_stats.html', context)

